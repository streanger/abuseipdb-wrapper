import json
import time
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from abuseipdb_wrapper.logger import RED_LEVEL, YELLOW_LEVEL

# colors are specific to html and xlsx so we dont import them from logger
GREEN = "#4cf58c"
YELLOW = "#f5cd4c"
RED = "#f54c4c"
BLUE = "#00ffff"
# ONION = "💀"  # 🧅
MARK_TRUE = "✔"
MARK_FALSE = "✘"
IS_TOR = "isTor"
ABUSE_CONFIDENCE_SCORE = "abuseConfidenceScore"
GREEN_XLSX = PatternFill(start_color=GREEN.lstrip('#'), end_color=GREEN.lstrip('#'), fill_type="solid")
YELLOW_XLSX = PatternFill(start_color=YELLOW.lstrip('#'), end_color=YELLOW.lstrip('#'), fill_type="solid")
RED_XLSX = PatternFill(start_color=RED.lstrip('#'), end_color=RED.lstrip('#'), fill_type="solid")
BLUE_XLSX = PatternFill(start_color=BLUE.lstrip('#'), end_color=BLUE.lstrip('#'), fill_type="solid")


def apply_css_style(value):
    """returns corresponding css style; it is not direct color"""
    if value >= RED_LEVEL:
        color = 'red'
    elif YELLOW_LEVEL <= value < RED_LEVEL:
        color = 'yellow'
    else:
        color = 'green'
    return color


def apply_color_xlsx(value):
    """returns xlsx color object"""
    if value >= RED_LEVEL:
        color = RED_XLSX
    elif YELLOW_LEVEL <= value < RED_LEVEL:
        color = YELLOW_XLSX
    else:
        color = GREEN_XLSX
    return color


def read_json(filename):
    """read json file to dict"""
    data = {}
    try:
        with open(filename, encoding="utf-8") as f:
            data = json.load(f)
    except FileNotFoundError:
        print("[x] FileNotFoundError: {}".format(filename))
    return data


def to_xlsx(data, hide=None):
    if hide is None:
        hide = []

    cambria_font = Font(name='Cambria', bold=True)
    calibri_font = Font(name='Calibri')
    center_alignment = Alignment(horizontal='center', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')
    thin_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

    # Create a workbook and select active worksheet
    wb = Workbook()
    ws = wb.active

    # Write headers with blue color
    header, *data = data
    header_map = {value: index for index, value in enumerate(header)}
    for hidden in hide:
        del header[header_map[hidden]]
    for col_index, value in enumerate(header, start=2):
        cell = ws.cell(row=1, column=col_index, value=value)
        cell.font = cambria_font
        cell.alignment = center_alignment
        cell.border = thin_border

    # Write data with color based on condition
    for row_index, row in enumerate(data, start=2):
        index_cell = ws.cell(row=row_index, column=1, value=row_index-1)  # Write index
        index_cell.font = cambria_font
        index_cell.alignment = center_alignment
        index_cell.border = thin_border
        row_color = apply_color_xlsx(row[header_map[ABUSE_CONFIDENCE_SCORE]])

        # remove hidden item(s)
        for hidden in hide:
            del row[header_map[hidden]]

        for col_index, col in enumerate(row, start=2):
            if type(col) is list:
                col = str(col)
            cell = ws.cell(row=row_index, column=col_index, value=col)
            cell_color = row_color
            cell.fill = cell_color
            cell.font = calibri_font
            cell.alignment = right_alignment
    return wb


def to_html(data, hide=None):
    """convert table with list of lists to html table

    data - data to show as html table
    hide - list of columns to hide
    code is modfied version of func from sets_matcher.py
    """
    if hide is None:
        hide = []

    # **** create body ****
    header, *table = data
    header = header.copy()
    tab = ' '*4
    table_body = ""
    header_map = {value: index for index, value in enumerate(header)}
    for row_index, row in enumerate(table, start=1):
        row_color = apply_css_style(row[header_map[ABUSE_CONFIDENCE_SCORE]])
        # if row[header_map[IS_TOR]]:
        #     row[header_map[IS_TOR]] = ONION

        # **** remove hidden item(s) ****
        for hidden in hide:
            del row[header_map[hidden]]

        # **** iterate cells ****
        cells = []
        cells.append(f"{tab*4}<td>{row_index}</td>\n")
        for index, column in enumerate(row):
            cell_style= ''
            if str(column).startswith('https://'):
                column = f"<a href=\"{column}\">url</a>"
            elif type(column) is bool:
                if column:
                    cell_style= ' class="marker"'
                    column = MARK_TRUE
                else:
                    column = MARK_FALSE
            cells.append(f"{tab*4}<td{cell_style}>{column}</td>\n")
        cells = ''.join(cells)
        row_style = f' class="{row_color}"'
        table_body += f"{tab*3}<tr{row_style}>\n{cells}{tab*3}</tr>\n"
    table_body = table_body.rstrip()

    # **** create head ****
    for hidden in hide:
        del header[header_map[hidden]]
    header.insert(0, 'Index')
    table_head = '\n'.join([f"{tab*3}<th><button>{column}</button></th>" for column in header])

    # INFO: this is non-breaking space: \00a0
    style = '''\
    .styled-table {
        border-collapse: collapse;
        margin-left: auto;
        margin-right: auto;
        font-size: 0.8em;
        font-family: cambria;
        min-width: 400px;
    }
    .styled-table thead tr {
        background-color: #0099d9;
        color: #ffffff;
    }
    .styled-table td {
        padding: 6px 9px;
        text-align: center;
    }
    .styled-table td:first-child {
        padding: 6px 9px;
        text-align: right;
    }
    .styled-table tbody tr {
        border-bottom: 1px solid #0099d9;
    }
    .styled-table th {
        padding: 0;
        text-align: center;
    }
    .styled-table th button {
        background-color: transparent;
        border: none;
        font: inherit;
        color: inherit;
        height: 100%;
        width: 100%;
        padding: 6px 9px;
        display: inline-block;
    }
    .styled-table th button::after {
        content: "\\00a0\\00a0";
        font-family: 'Courier New', Courier, monospace
    }
    .styled-table th button[direction="ascending"]::after {
        content: "\\00a0▲";
    }
    .styled-table th button[direction="descending"]::after {
        content: "\\00a0▼";
    }
    .red {background-color: #f54c4c;}
    .yellow {background-color: #f5cd4c;}
    .green {background-color: #4cf58c;}
    .marker {
        background-color: #ffffff99;
        border-radius: 30px;
    }'''

# INFO: for both IPv4 and IPv6
# function compareIPs(a, b) {
#     const aIsIPv6 = a.includes(':');
#     const bIsIPv6 = b.includes(':');
#     // If one is IPv4 and the other is IPv6, sort IPv4 first
#     if (aIsIPv6 !== bIsIPv6) return aIsIPv6 ? 1 : -1;
#     const aParts = a.split(aIsIPv6 ? ':' : '.').map(part => parseInt(part, aIsIPv6 ? 16 : 10));
#     const bParts = b.split(bIsIPv6 ? ':' : '.').map(part => parseInt(part, bIsIPv6 ? 16 : 10));
#     // Compare each part of the IP
#     for (let i = 0; i < aParts.length; i++) {
#         if ((aParts[i] || 0) > (bParts[i] || 0)) return 1;
#         if ((aParts[i] || 0) < (bParts[i] || 0)) return -1;
#     }
#     return 0;
# }

    script = """\
function main() {
    var table = document.getElementsByTagName("table")[0];
    var header = table.getElementsByTagName("tr")[0];
    var headers = header.getElementsByTagName("th");
    for (var i = 0; i < headers.length; i++) {
        var btn = headers[i].getElementsByTagName("button")[0];
        btn.setAttribute("onclick", `table_sorter(${i})`);
    }
}

function compareIPs(a, b) {
    const aParts = a.split('.').map(part => +part);
    const bParts = b.split('.').map(part => +part);
    for (let i = 0; i < aParts.length; i++) {
        if (aParts[i] > bParts[i]) return 1;
        if (aParts[i] < bParts[i]) return -1;
    }
    return 0;
}

function table_sorter(column) {
    var table = document.getElementsByTagName("table")[0];
    var tableBody = table.getElementsByTagName("tbody")[0];
    var columnButton = table.getElementsByTagName("tr")[0].getElementsByTagName("th")[column].getElementsByTagName("button")[0];
    var columnName = columnButton.textContent;
    var direction = columnButton.getAttribute("direction");
    if (direction == "ascending") {
        direction = "descending";
    } else {
        direction = "ascending";
    }
    var rows = Array.from(table.getElementsByTagName("tr")).slice(1);
    if (columnName == "ipAddress") {
        rows.sort(function(a, b) {
            var x = a.getElementsByTagName("td")[column].textContent.toLowerCase();
            var y = b.getElementsByTagName("td")[column].textContent.toLowerCase();
            if (direction === "ascending") {
                return compareIPs(x, y) || x.localeCompare(y);
            } else {
                return compareIPs(y, x) || y.localeCompare(x);
            }
        });
    }
    else {
        rows.sort(function(a, b) {
            var x = a.getElementsByTagName("td")[column].textContent.toLowerCase();
            var y = b.getElementsByTagName("td")[column].textContent.toLowerCase();
            if (direction === "ascending") {
                return x - y || x.localeCompare(y);
            } else {
                return y - x || y.localeCompare(x);
            }
        });
    }
    rows.forEach(function(row) {
        tableBody.appendChild(row);
    });

    // show direction using arrow icon
    var header = table.getElementsByTagName("tr")[0];
    var headers = header.getElementsByTagName("th");
    for (var i = 0; i < headers.length; i++) {
        var btn = headers[i].getElementsByTagName("button")[0];
        if (i == column) {
            btn.setAttribute("direction", direction);
        } else {
            btn.setAttribute("direction", "");
        }
    }
}"""

    template = f"""\
<html>
<head>
    <title>abuse</title>
    <meta charset="utf-8">
    <style>
{style}
    </style>
    <script>
{script}
    </script>
</head>
<body onload=main()>
    <table class="styled-table">
        <thead>
            <tr>
{table_head}
            </tr>
        </thead>
        <tbody>
{table_body}
        </tbody>
    </table>
</body>
</html>\
"""
    return template


if __name__ == '__main__':
    db = read_json('tests/test_home/abuse.json')
    matched = list(db.values())
    temporary_columns = list(matched[0].keys()).copy()
    matched = [[row[column] for column in temporary_columns] for row in matched]
    matched.insert(0, temporary_columns)

    html = to_html(data=matched)
    xlsx = to_xlsx(data=matched)

    # FIXED: use style instead of background-color all the time (3 styles for red, green, yellow)
    # FIXED: first column index
    # REFUSED: pseudo dataframe to match columns
    # FIXED: true ip sorter for html
    # FIXED: make sorter arrow not to flow
    # REFUSED: switch with night mode
    # REFUSED: take colors from terminal
    # REFUSED: clickable url in excel
    # REFUSED: onion ascii in html
    # FIXED: date should be in iso format
    # REFUSED: add index in the front (csv export)
    # REFUSED: url sorter should use value (in html)
    # REFUSED: index in html maybe should stay in place?
    # REFUSED: minify html template (js, etc) -> https://github.com/ndparker/rjsmin
    # FIXED: apply style to tr instead of td
    # REFUSED: https://github.com/dompdf/dompdf/issues/2301 -> border-collapse: collapse; vs border-collapse: separate;
