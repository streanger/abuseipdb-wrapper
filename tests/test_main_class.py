import json
from pathlib import Path
from openpyxl import load_workbook
import pytest
from pytest_httpserver import HTTPServer
from rich import print
from abuseipdb_wrapper import AbuseIPDB


def read_json(filename):
    """read json file to dict"""
    data = {}
    try:
        with open(filename, encoding="utf-8") as f:
            data = json.load(f)
    except FileNotFoundError:
        print("[x] FileNotFoundError: {}".format(filename))
    return data


def read_xlsx(filename):
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    data = [list(row) for row in sheet.iter_rows(values_only=True)]
    return data


def test_main_class_instance():
    api_key = 'XXXXXX'
    input_ips = ['1.1.1.1', '2.2.2.2', '3.3.3']
    asserted_ips = ['1.1.1.1', '2.2.2.2']
    abuse = AbuseIPDB(api_key=api_key, verbose=True, ip_list=input_ips)
    assert abuse
    assert abuse.config['verbose'] == True
    assert abuse._ip_list == asserted_ips


def test_api_key_not_set():
    abuse = AbuseIPDB()
    with pytest.raises(Exception) as err:
        abuse.check_ip('1.2.3.4')
    assert str(err.value) == "API key not set"


def test_check_ip(httpserver: HTTPServer):
    raw_responses = {
        '1.1.1.1': '{"data":{"ipAddress":"1.1.1.1","isPublic":true,"ipVersion":4,"isWhitelisted":true,"abuseConfidenceScore":0,"countryCode":"US","usageType":"Content Delivery Network","isp":"APNIC and CloudFlare DNS Resolver Project","domain":"cloudflare.com","hostnames":["one.one.one.one"],"isTor":false,"totalReports":43,"numDistinctUsers":25,"lastReportedAt":"2024-03-04T00:00:20+00:00"}}',
        '2.2.2.2': '{"data":{"ipAddress":"2.2.2.2","isPublic":true,"ipVersion":4,"isWhitelisted":false,"abuseConfidenceScore":0,"countryCode":"FR","usageType":null,"isp":"Orange S.A.","domain":"orange.com","hostnames":[],"isTor":false,"totalReports":3,"numDistinctUsers":3,"lastReportedAt":"2024-02-26T10:40:58+00:00"}}',
        '3.3.3.3': '{"data":{"ipAddress":"3.3.3.3","isPublic":true,"ipVersion":4,"isWhitelisted":false,"abuseConfidenceScore":21,"countryCode":"US","usageType":"Data Center\\/Web Hosting\\/Transit","isp":"Amazon Technologies Inc.","domain":"amazon.com","hostnames":[],"isTor":false,"totalReports":5,"numDistinctUsers":4,"lastReportedAt":"2024-03-02T09:32:41+00:00"}}',
        '4.4.4.4': '{"data":{"ipAddress":"4.4.4.4","isPublic":true,"ipVersion":4,"isWhitelisted":false,"abuseConfidenceScore":0,"countryCode":"US","usageType":"Fixed Line ISP","isp":"Level 3 Communications Inc.","domain":"level3.com","hostnames":[],"isTor":false,"totalReports":0,"numDistinctUsers":0,"lastReportedAt":"2024-01-13T20:50:19+00:00"}}',
        '1.2.3.4': '{"data":{"ipAddress":"1.2.3.4","isPublic":true,"ipVersion":4,"isWhitelisted":false,"abuseConfidenceScore":9,"countryCode":"AU","usageType":"Data Center\\/Web Hosting\\/Transit","isp":"APNIC Pty Ltd","domain":"apnic.net","hostnames":[],"isTor":false,"totalReports":8,"numDistinctUsers":3,"lastReportedAt":"2024-03-05T13:20:17+00:00"}}',
    }
    # **** setup ****
    db_file = None
    config_file = None
    api_key = 'XXXXXX'
    abuse = AbuseIPDB(api_key=api_key, db_file=db_file, config_file=config_file)
    abuse._abuse_domain = f'http://localhost:{httpserver.port}'
    verbose = False
    endpoint = "/api/v2/check"
    max_age_in_days = "90"

    IPs = ['1.1.1.1', '2.2.2.2', '3.3.3.3', '4.4.4.4', '1.2.3.4']
    for IP_TO_CHECK in IPs:
        # **** prepare server response
        if verbose:
            querystring = {
                "ipAddress": IP_TO_CHECK,
                "maxAgeInDays": max_age_in_days,
                "verbose": verbose,
            }
        else:
            querystring = {
                "ipAddress": IP_TO_CHECK,
                "maxAgeInDays": max_age_in_days,
            }
        headers = {"Accept": "application/json", "Key": api_key}
        response = raw_responses[IP_TO_CHECK]
        httpserver.expect_request(endpoint, headers=headers, query_string=querystring).respond_with_data(response)

        # **** check_ip ****
        result = abuse.check_ip(IP_TO_CHECK)
        del result['date']
        del result['url']
        assert result == json.loads(raw_responses[IP_TO_CHECK])['data']

    # **** test check() ****
    abuse.add_ip_list(IPs)
    abuse.check()


def test_add_ip_list():
    api_key = 'XXXXXX'
    abuse = AbuseIPDB(api_key=api_key)
    ip_list = [
        '1.1.1.1',
        '2.2.2.2',
        '3.3.3.3',
        '4.4.4.4',
    ]
    abuse.add_ip_list(ip_list=ip_list)
    assert abuse._ip_list == ip_list
    abuse.clear_ip_list()
    assert abuse._ip_list == []


def test_db_interaction():
    CORRECT_TABLE = """\
┏━━━━┳━━━━━━━━━━━┳━━━━━━━━━━━━━━━━━━━━━━┳━━━━━━━━━━━━━┳━━━━━━━━━━━━━━━━┳━━━━━━━━━━━━━━┳━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┳━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓
┃ No ┃ ipAddress ┃ abuseConfidenceScore ┃ countryCode ┃ domain         ┃ totalReports ┃ url                                     ┃ usageType                       ┃
┡━━━━╇━━━━━━━━━━━╇━━━━━━━━━━━━━━━━━━━━━━╇━━━━━━━━━━━━━╇━━━━━━━━━━━━━━━━╇━━━━━━━━━━━━━━╇━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━╇━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┩
│ 1  │ 1.1.1.1   │ 0                    │ US          │ cloudflare.com │ 42           │ https://www.abuseipdb.com/check/1.1.1.1 │ Content Delivery Network        │
│ 2  │ 2.2.2.2   │ 0                    │ FR          │ orange.com     │ 3            │ https://www.abuseipdb.com/check/2.2.2.2 │ None                            │
│ 3  │ 3.3.3.3   │ 22                   │ US          │ amazon.com     │ 5            │ https://www.abuseipdb.com/check/3.3.3.3 │ Data Center/Web Hosting/Transit │
│ 4  │ 4.4.4.4   │ 0                    │ US          │ level3.com     │ 0            │ https://www.abuseipdb.com/check/4.4.4.4 │ Fixed Line ISP                  │
│ 5  │ 5.5.5.5   │ 0                    │ DE          │ telefonica.de  │ 1            │ https://www.abuseipdb.com/check/5.5.5.5 │ None                            │
│ 6  │ 6.6.6.6   │ 0                    │ US          │ army.mil       │ 1            │ https://www.abuseipdb.com/check/6.6.6.6 │ Military                        │
│ 7  │ 7.7.7.7   │ 0                    │ US          │ nic.mil        │ 0            │ https://www.abuseipdb.com/check/7.7.7.7 │ Military                        │
│ 8  │ 8.8.8.8   │ 0                    │ US          │ google.com     │ 53           │ https://www.abuseipdb.com/check/8.8.8.8 │ Data Center/Web Hosting/Transit │
└────┴───────────┴──────────────────────┴─────────────┴────────────────┴──────────────┴─────────────────────────────────────────┴─────────────────────────────────┘
"""
    # **** setup ****
    api_key = 'XXXXXX'
    db_file = 'tests/test_home/abuse.json'
    config_file = 'tests/test_home/config.json'
    abuse = AbuseIPDB(api_key=api_key, db_file=db_file, config_file=config_file)

    # **** check table ****
    # print(f'terminal size: {os.get_terminal_size()}')
    # print(abuse)
    # abuse.show()
    # table = str(abuse)
    # table2 = abuse.__str__(abuse)
    # view = abuse._create_view(matched_only=False, table_view=True)
    # plain_text = abuse._log_table(view).plain
    # assert table == CORRECT_TABLE

    # **** check vertical view ****

    # **** check other methods ****
    correct_db = read_json(db_file)
    db = abuse.get_db()
    assert db == correct_db

    # **** set/get columns ****
    new_columns = [
        "abuseConfidenceScore",
        "countryCode",
        "domain",
        "hostnames",
        "ipAddress",
        "ipVersion",
        "isPublic",
        "isWhitelisted",
        "isp",
        "lastReportedAt",
        "numDistinctUsers",
        "totalReports",
        "usageType",
        "url",  # additional (abuseipdb related url)
        "date",  # additional (date of request)
        "isTor",
    ]
    abuse.set_columns(new_columns)
    assert abuse.config['columns'] == new_columns
    assert abuse.config['columns'] == abuse.get_all_columns()
    abuse.set_columns(abuse.get_init_columns())


def test_exports():
    # **** setup ****
    api_key = 'XXXXXX'
    db_file = 'tests/test_home/abuse.json'
    config_file = 'tests/test_home/config.json'
    abuse = AbuseIPDB(api_key=api_key, db_file=db_file, config_file=config_file)

    # **** test exports ****
    test_csv, new_csv = 'tests/test_home/test.csv', 'tests/test_home/out.csv'
    abuse.export_csv(path=new_csv)
    assert Path(new_csv).read_bytes() == Path(test_csv).read_bytes()

    test_html, new_html = 'tests/test_home/test.html', 'tests/test_home/out.html'
    abuse.export_html_styled(path=new_html)
    assert Path(new_html).read_bytes() == Path(test_html).read_bytes()

    test_xlsx, new_xlsx, = 'tests/test_home/test.xlsx', 'tests/test_home/out.xlsx'
    abuse.export_xlsx_styled(path=new_xlsx)
    assert read_xlsx(new_xlsx) == read_xlsx(test_xlsx)

    test_md, new_md, = 'tests/test_home/test.md', 'tests/test_home/out.md'
    abuse.export_md(path=new_md)
    assert Path(new_md).read_bytes() == Path(test_md).read_bytes()
